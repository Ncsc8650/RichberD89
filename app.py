from http import cookies
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse
import datetime as dt
import hashlib
import io
import json
import mimetypes
import os
import secrets
import sqlite3
import sys
import traceback

from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "richber.sqlite3"
ADMIN_USER = "rich89"
ADMIN_PASSWORD = "89rich"
SESSIONS = set()

MONEY_CHARM = {"22", "24", "42", "26", "62", "36", "63", "46", "64", "56", "65", "66", "69", "96"}
WISDOM = {"14", "41", "15", "51", "16", "61", "45", "54", "55", "56", "65", "59", "95"}
FRIENDS = {"15", "51", "24", "42", "36", "63", "78", "87"}
LOVE = {"22", "26", "62", "24", "42", "46", "64", "65", "56", "66"}
BAD_PAIRS = {"00", "01", "10", "02", "20", "03", "30", "06", "60", "08", "18", "81", "21", "12", "25", "52", "27", "72", "31", "13", "37", "73", "38", "83", "43", "34", "47", "74", "48", "84", "50", "05"}
GOOD_ENDINGS = {"5", "9"}
NEGATIVE_ENDINGS = {"0", "1", "3", "7", "8"}
STATUS_ORDER = {"จองแล้ว": 1, "รอขาย": 2, "จำหน่ายแล้ว": 3, "จำหน่ายไปแล้ว": 3}
PAIR_MEANINGS = {
    "14": "พูดดี มีความรู้ วิชาการและการสื่อสารเด่น", "41": "พูดดี มีปัญญา วิชาการเด่น",
    "15": "ปัญญาละเอียด น่าเชื่อถือ ผู้ใหญ่เมตตา", "51": "ปัญญาผู้ใหญ่ น่าเชื่อถือ ผู้ใหญ่เมตตา",
    "16": "บริหารเงินดี คิดละเอียด วางแผนเก่ง", "61": "วางแผนเก่ง เก็บเงินดี คิดละเอียด",
    "22": "พูดเก่ง มีเสน่ห์ อ่อนโยน", "24": "เมตตามหานิยม คนสนับสนุน การเงินดี", "42": "เมตตาค้าขาย คนช่วยเหลือ เงินเข้าตลอด",
    "26": "พูดดี ได้ทรัพย์จากการพูด โรแมนติก", "62": "เสน่ห์ อ่อนโยน ได้ทรัพย์จากคำพูด",
    "36": "คู่มิตร เสน่ห์ดี หาเงินคล่อง", "63": "คู่มิตร เสน่ห์ ทรัพย์ หาเงินคล่อง",
    "45": "สติปัญญา เหตุผล ความน่าเชื่อถือ", "54": "ปัญญา เหตุผล การเรียน งานที่ปรึกษา", "55": "ธรรมะ เหตุผล รอบคอบ",
    "56": "ทรัพย์และปัญญา คุมการเงินและการตัดสินใจ", "65": "ทรัพย์เสน่ห์และสติปัญญา เป็นคู่เด่นมาก",
    "59": "สิ่งศักดิ์สิทธิ์ ปัญญา โชค", "95": "ปัญญา สิ่งศักดิ์สิทธิ์ โชคและความคุ้มครอง",
    "66": "เสน่ห์ ความสบาย ทรัพย์และความนิยม", "69": "เสน่ห์ เงิน ความงาม ศิลปะ", "96": "เสน่ห์ ความงาม ทรัพย์ ศิลปะ",
    "78": "คู่มิตรงานใหญ่ ความเสี่ยง หุ้น โครงการใหญ่", "87": "คู่มิตรงานใหญ่ ความเสี่ยงและการสนับสนุน",
    "00": "โลกส่วนตัวสูง เก็บความเครียด ควรระวังสุขภาพ", "03": "ใจร้อน รวดเร็ว ชอบลุย", "30": "ใจร้อน รวดเร็ว ชอบลุย",
    "06": "มีเสน่ห์แต่ใจอ่อน ใช้เงินง่าย ระวังเงินรั่ว", "60": "มีเสน่ห์แต่ใจอ่อน ใช้เงินง่าย ระวังเงินรั่ว",
    "27": "คู่หนี้สิน ระวังภาระการเงิน", "37": "อุปสรรคและแรงกดดันสูง", "38": "กล้าได้กล้าเสีย ตัดสินใจไว",
    "47": "ลุยงานหนัก พูดตรงแรง", "74": "ลุยงานหนัก พูดตรงแรง", "50": "คิดมาก ฟุ้งซ่าน ระวังความเครียด", "05": "คิดลึก ระวังคิดมากเกินไป",
}
CAREER_GROUPS = {
    "ราชการ/ผู้ใหญ่สนับสนุน": {"14", "41", "15", "51", "45", "54", "19", "91", "35", "53"},
    "ค้าขาย/เซลล์/นักลงทุน": {"22", "24", "42", "26", "62", "44", "46", "64", "56", "65", "66", "29", "92", "36", "63"},
    "สื่อสาร/ประชาสัมพันธ์": {"24", "42", "44", "46", "64", "47", "74"},
    "การศึกษา/ธรรมะ/ปัญญา": {"14", "41", "15", "51", "45", "54", "55", "59", "95", "99"},
    "ไฮเทค/ออนไลน์": {"19", "91", "49", "94", "59", "95", "99", "79", "97"},
    "ศิลปะ/ความงาม/บันเทิง": {"29", "92", "36", "63", "66", "69", "96", "24", "42"},
    "งานเสี่ยง/หุ้น/โครงการใหญ่": {"78", "87", "89", "98", "79", "97"},
}


def db():
    DATA_DIR.mkdir(exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def normalize_phone(value):
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def phone_pairs(phone):
    d = normalize_phone(phone)
    return [d[i:i + 2] for i in range(max(0, len(d) - 1))]


def analyze_phone(phone):
    d = normalize_phone(phone)
    pairs = phone_pairs(d)
    score = 50
    summary, highlights, warnings = [], [], []
    groups = {"ทรัพย์/เสน่ห์": sorted(set(pairs) & MONEY_CHARM), "สติปัญญา": sorted(set(pairs) & WISDOM), "คู่มิตร": sorted(set(pairs) & FRIENDS), "ความรักสดใส": sorted(set(pairs) & LOVE)}
    score += 7 * len(groups["ทรัพย์/เสน่ห์"]) + 7 * len(groups["สติปัญญา"]) + 5 * len(groups["คู่มิตร"]) + 3 * len(groups["ความรักสดใส"])
    bad = sorted(set(pairs) & BAD_PAIRS)
    score -= 6 * len(bad)
    if bad:
        warnings.append("มีเลขคู่ที่ควรใช้ด้วยความระวัง: " + ", ".join(bad))
    if d.endswith("65"):
        score += 15
        highlights.append("ลงท้าย 65 เด่นมาก: ทรัพย์ เสน่ห์ และปัญญา")
    elif d[-1:] in GOOD_ENDINGS:
        score += 8
        highlights.append(f"ลงท้ายด้วย {d[-1]} ช่วยเสริมสติหรือความคุ้มครอง")
    elif d[-1:] in NEGATIVE_ENDINGS:
        score -= 8
        warnings.append(f"ลงท้ายด้วย {d[-1]} ควรระวังอารมณ์ ความเครียด หรือเงินรั่ว")
    if "0" in d[3:]:
        score -= 4
        warnings.append("มีเลข 0 หลังตำแหน่งต้นเบอร์ ควรระวังปัญหาซ่อนเร้น")
    if d.count("8") > 1:
        score -= 4
        warnings.append("มีเลข 8 มากกว่า 1 ตัว เหมาะงานเสี่ยงแต่ควรคุมอารมณ์และการเงิน")
    if d.count("5") or d.count("9"):
        score += 4
        highlights.append("มีเลข 5 หรือ 9 ช่วยคุมพลังแรงของเบอร์")
    careers = []
    for title, vals in CAREER_GROUPS.items():
        found = sorted(set(pairs) & vals)
        if found:
            careers.append({"title": title, "pairs": found})
            score += min(5, len(found) * 2)
    for title, vals in groups.items():
        if vals:
            summary.append(f"{title}: {', '.join(vals)}")
    if not summary:
        summary.append("ยังไม่พบกลุ่มเลขเด่นตามสูตรหลัก")
    score = max(0, min(100, score))
    level = "ดีมาก" if score >= 85 else "ดี" if score >= 70 else "กลาง"
    return {"score": score, "level": level, "pairs": pairs, "summary": summary, "highlights": highlights, "warnings": warnings, "career_matches": careers[:5], "pair_details": [{"pair": p, "meaning": PAIR_MEANINGS.get(p, "เลขคู่นี้ใช้ประกอบคะแนนจากกลุ่มเลข")} for p in pairs]}


def init_db():
    with db() as conn:
        conn.execute("CREATE TABLE IF NOT EXISTS numbers (id INTEGER PRIMARY KEY AUTOINCREMENT, sequence_no INTEGER, phone TEXT UNIQUE NOT NULL, sale_price REAL DEFAULT 0, wholesale_price REAL DEFAULT 0, network TEXT DEFAULT '', status TEXT DEFAULT 'รอขาย', expiry_date TEXT DEFAULT '', created_at TEXT DEFAULT CURRENT_TIMESTAMP, updated_at TEXT DEFAULT CURRENT_TIMESTAMP)")
        conn.execute("CREATE TABLE IF NOT EXISTS statuses (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL, sort_order INTEGER DEFAULT 0, created_at TEXT DEFAULT CURRENT_TIMESTAMP, updated_at TEXT DEFAULT CURRENT_TIMESTAMP)")
        conn.execute("CREATE TABLE IF NOT EXISTS networks (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL, sort_order INTEGER DEFAULT 0, created_at TEXT DEFAULT CURRENT_TIMESTAMP, updated_at TEXT DEFAULT CURRENT_TIMESTAMP)")
        conn.execute("CREATE TABLE IF NOT EXISTS import_logs (id INTEGER PRIMARY KEY AUTOINCREMENT, filename TEXT DEFAULT '', total_rows INTEGER DEFAULT 0, imported_rows INTEGER DEFAULT 0, duplicate_rows INTEGER DEFAULT 0, error_rows INTEGER DEFAULT 0, replaced_all INTEGER DEFAULT 0, details TEXT DEFAULT '[]', created_at TEXT DEFAULT CURRENT_TIMESTAMP)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_numbers_phone ON numbers(phone)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_numbers_status ON numbers(status)")
        for i, name in enumerate(["จองแล้ว", "รอขาย", "จำหน่ายไปแล้ว"], 1):
            conn.execute("INSERT OR IGNORE INTO statuses(name, sort_order) VALUES(?, ?)", [name, i])


def parse_date(value):
    if not value:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def ensure_master(conn, table, name):
    clean = str(name or "").strip()
    if clean:
        conn.execute(f"INSERT OR IGNORE INTO {table}(name) VALUES(?)", [clean])


def list_master(table):
    with db() as conn:
        rows = [dict(r) for r in conn.execute(f"SELECT * FROM {table} ORDER BY sort_order, name")]
    if table == "statuses":
        rows.sort(key=lambda r: (STATUS_ORDER.get(r["name"], 99), r.get("sort_order") or 0, r["name"]))
    return rows


def save_master(table, data):
    name = str(data.get("name") or "").strip()
    if not name:
        raise ValueError("กรุณาระบุชื่อ")
    payload = {"id": data.get("id") or None, "name": name, "sort_order": int(data.get("sort_order") or 0)}
    with db() as conn:
        if payload["id"]:
            conn.execute(f"UPDATE {table} SET name=:name, sort_order=:sort_order, updated_at=CURRENT_TIMESTAMP WHERE id=:id", payload)
        else:
            conn.execute(f"INSERT INTO {table}(name, sort_order) VALUES(:name, :sort_order)", payload)


def delete_master(table, item_id):
    field = "status" if table == "statuses" else "network"
    with db() as conn:
        item = conn.execute(f"SELECT name FROM {table} WHERE id = ?", [item_id]).fetchone()
        if not item:
            raise ValueError("ไม่พบข้อมูล")
        used = conn.execute(f"SELECT COUNT(*) FROM numbers WHERE {field} = ?", [item["name"]]).fetchone()[0]
        if used:
            raise ValueError(f"ไม่สามารถลบได้ เพราะมีเบอร์ใช้งานอยู่ {used} รายการ")
        conn.execute(f"DELETE FROM {table} WHERE id = ?", [item_id])


def import_excel(path, filename="", replace_all=False):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"total_rows": 0, "imported_rows": 0, "duplicate_rows": 0, "error_rows": 0, "details": []}
    headers = [str(h or "").strip() for h in rows[0]]
    idx = {name: i for i, name in enumerate(headers)}
    aliases = {"sequence_no": ["ลำดับ", "no", "seq"], "phone": ["หมายเลขโทรศัพท์", "เบอร์โทรศัพท์", "phone"], "sale_price": ["ราคาขาย", "sale_price"], "wholesale_price": ["ราคาส่ง", "wholesale_price"], "network": ["เครือข่าย", "network"], "status": ["สถานะ", "status"], "expiry_date": ["วันหมดอายุ", "expiry_date"]}

    def pick(row, key, default=""):
        for name in aliases[key]:
            if name in idx and idx[name] < len(row):
                return row[idx[name]]
        return default

    imported = duplicates = errors = 0
    details, seen = [], set()
    with db() as conn:
        if replace_all:
            conn.execute("DELETE FROM numbers")
        for row_no, row in enumerate(rows[1:], 2):
            try:
                phone = normalize_phone(pick(row, "phone"))
                if not phone:
                    raise ValueError("ไม่มีหมายเลขโทรศัพท์")
                if phone in seen:
                    duplicates += 1
                    details.append({"row": row_no, "phone": phone, "status": "duplicate", "message": "เบอร์ซ้ำในไฟล์ Excel"})
                    continue
                seen.add(phone)
                if conn.execute("SELECT id FROM numbers WHERE phone = ?", [phone]).fetchone():
                    duplicates += 1
                    details.append({"row": row_no, "phone": phone, "status": "duplicate", "message": "เบอร์ซ้ำกับข้อมูลเดิมในฐานข้อมูล"})
                    continue
                values = {"sequence_no": int(pick(row, "sequence_no") or 0), "phone": phone, "sale_price": float(pick(row, "sale_price") or 0), "wholesale_price": float(pick(row, "wholesale_price") or 0), "network": str(pick(row, "network") or "").strip(), "status": str(pick(row, "status") or "รอขาย").strip(), "expiry_date": parse_date(pick(row, "expiry_date"))}
                ensure_master(conn, "networks", values["network"])
                ensure_master(conn, "statuses", values["status"])
                conn.execute("INSERT INTO numbers(sequence_no, phone, sale_price, wholesale_price, network, status, expiry_date) VALUES(:sequence_no, :phone, :sale_price, :wholesale_price, :network, :status, :expiry_date)", values)
                imported += 1
            except Exception as exc:
                errors += 1
                details.append({"row": row_no, "phone": normalize_phone(pick(row, "phone")), "status": "error", "message": str(exc)})
        conn.execute("INSERT INTO import_logs(filename, total_rows, imported_rows, duplicate_rows, error_rows, replaced_all, details) VALUES(?, ?, ?, ?, ?, ?, ?)", [filename, max(0, len(rows) - 1), imported, duplicates, errors, 1 if replace_all else 0, json.dumps(details, ensure_ascii=False)])
        log_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return {"total_rows": max(0, len(rows) - 1), "imported_rows": imported, "duplicate_rows": duplicates, "error_rows": errors, "details": details, "replaced_all": bool(replace_all), "log_id": log_id}


def query_numbers(params, public_only=False, randomize=False, limit=20, offset=0, admin_order=False):
    where, args = [], []
    if public_only:
        where.append("COALESCE(status, '') NOT IN (?, ?)")
        args.extend(["จำหน่ายแล้ว", "จำหน่ายไปแล้ว"])
    search = (params.get("q") or [""])[0].strip()
    if search:
        like = f"%{search}%"
        where.append("(phone LIKE ? OR network LIKE ? OR status LIKE ?)")
        args.extend([like, like, like])
    for key, column in (("network", "network"), ("status", "status")):
        value = (params.get(key) or [""])[0].strip()
        if value:
            where.append(f"{column} = ?")
            args.append(value)
    for key, op in (("min_price", ">="), ("max_price", "<=")):
        value = (params.get(key) or [""])[0].strip()
        if value:
            where.append(f"sale_price {op} ?")
            args.append(float(value))
    clause = " WHERE " + " AND ".join(where) if where else ""
    if randomize:
        order = " ORDER BY RANDOM()"
    elif admin_order:
        order = " ORDER BY CASE status WHEN 'จองแล้ว' THEN 1 WHEN 'รอขาย' THEN 2 WHEN 'จำหน่ายแล้ว' THEN 3 WHEN 'จำหน่ายไปแล้ว' THEN 3 ELSE 4 END, sequence_no ASC, id ASC"
    else:
        order = " ORDER BY id DESC"
    with db() as conn:
        rows = [dict(r) for r in conn.execute(f"SELECT * FROM numbers{clause}{order} LIMIT ? OFFSET ?", args + [limit, offset])]
        total = conn.execute(f"SELECT COUNT(*) FROM numbers{clause}", args).fetchone()[0]
    for row in rows:
        row["analysis"] = analyze_phone(row["phone"])
    return {"items": rows, "total": total}


def json_response(handler, payload, status=200):
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def file_response(handler, data, filename, content_type):
    handler.send_response(200)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Content-Disposition", f'attachment; filename="{filename}"')
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def pdf_font_name():
    candidates = [BASE_DIR / "Fonts" / "tahoma.ttf", BASE_DIR / "fonts" / "NotoSansThai-Regular.ttf", Path(os.environ.get("WINDIR", "C:\\Windows")) / "Fonts" / "tahoma.ttf"]
    for path in candidates:
        if path.exists():
            pdfmetrics.registerFont(TTFont("ThaiPDF", str(path)))
            return "ThaiPDF"
    return "Helvetica"


class App(BaseHTTPRequestHandler):
    server_version = "RichBerD89/1.0"

    def log_message(self, fmt, *args):
        sys.stderr.write("[%s] %s\n" % (self.log_date_time_string(), fmt % args))

    def is_admin(self):
        jar = cookies.SimpleCookie(self.headers.get("Cookie", ""))
        sid = jar.get("session")
        return bool(sid and sid.value in SESSIONS)

    def read_body(self):
        return self.rfile.read(int(self.headers.get("Content-Length", "0") or 0))

    def read_json(self):
        return json.loads(self.read_body().decode("utf-8") or "{}")

    def send_index(self):
        data = (STATIC_DIR / "index.html").read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        parsed = urlparse(self.path)
        path, params = parsed.path, parse_qs(parsed.query)
        if path == "/":
            return self.send_index()
        if path.startswith("/static/"):
            target = (BASE_DIR / path.lstrip("/")).resolve()
            if not str(target).startswith(str(STATIC_DIR.resolve())) or not target.exists():
                return self.send_error(404)
            data = target.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", mimetypes.guess_type(target.name)[0] or "application/octet-stream")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if path == "/api/numbers":
            page = max(1, int((params.get("page") or ["1"])[0] or 1))
            return json_response(self, query_numbers(params, True, (params.get("random") or ["1"])[0] != "0", 20, (page - 1) * 20))
        if path.startswith("/api/numbers/"):
            with db() as conn:
                row = conn.execute("SELECT * FROM numbers WHERE id = ?", [path.rsplit("/", 1)[-1]]).fetchone()
            if not row:
                return json_response(self, {"error": "ไม่พบข้อมูล"}, 404)
            item = dict(row)
            item["analysis"] = analyze_phone(item["phone"])
            return json_response(self, item)
        if path == "/api/meta":
            return json_response(self, {"networks": list_master("networks"), "statuses": list_master("statuses"), "admin": self.is_admin()})
        if path == "/api/admin/list" and self.is_admin():
            page = max(1, int((params.get("page") or ["1"])[0] or 1))
            return json_response(self, query_numbers(params, limit=50, offset=(page - 1) * 50, admin_order=True))
        if path in {"/api/admin/statuses", "/api/admin/networks"} and self.is_admin():
            return json_response(self, {"items": list_master("statuses" if "statuses" in path else "networks")})
        if path == "/api/admin/import-logs" and self.is_admin():
            with db() as conn:
                logs = [dict(r) for r in conn.execute("SELECT id, filename, total_rows, imported_rows, duplicate_rows, error_rows, replaced_all, created_at FROM import_logs ORDER BY id DESC LIMIT 30")]
            return json_response(self, {"items": logs})
        if path.startswith("/api/admin/import-logs/") and self.is_admin():
            with db() as conn:
                row = conn.execute("SELECT * FROM import_logs WHERE id = ?", [path.rsplit("/", 1)[-1]]).fetchone()
            if not row:
                return json_response(self, {"error": "ไม่พบ log"}, 404)
            item = dict(row)
            item["details"] = json.loads(item.get("details") or "[]")
            return json_response(self, item)
        if path == "/api/admin/export.xlsx" and self.is_admin():
            return self.export_xlsx(params)
        if path == "/api/admin/export.pdf" and self.is_admin():
            return self.export_pdf(params)
        if path == "/api/admin/template.xlsx" and self.is_admin():
            return self.template_xlsx()
        if path.startswith("/api/admin/"):
            return json_response(self, {"error": "กรุณา login"}, 401)
        return self.send_index()

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/login":
            data = self.read_json()
            ok = data.get("user") == ADMIN_USER and secrets.compare_digest(hashlib.sha256((data.get("password") or "").encode()).hexdigest(), hashlib.sha256(ADMIN_PASSWORD.encode()).hexdigest())
            if not ok:
                return json_response(self, {"error": "userid หรือ password ไม่ถูกต้อง"}, 403)
            sid = secrets.token_urlsafe(32)
            SESSIONS.add(sid)
            body = b'{"ok":true}'
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Set-Cookie", f"session={sid}; HttpOnly; SameSite=Lax; Path=/")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if parsed.path == "/api/logout":
            self.send_response(204)
            self.send_header("Set-Cookie", "session=; Max-Age=0; Path=/")
            self.end_headers()
            return
        if not self.is_admin():
            return json_response(self, {"error": "กรุณา login"}, 401)
        if parsed.path == "/api/admin/save":
            data = self.read_json()
            phone = normalize_phone(data.get("phone"))
            if not phone:
                return json_response(self, {"error": "กรุณาระบุหมายเลขโทรศัพท์"}, 400)
            payload = {"id": data.get("id") or None, "sequence_no": int(data.get("sequence_no") or 0), "phone": phone, "sale_price": float(data.get("sale_price") or 0), "wholesale_price": float(data.get("wholesale_price") or 0), "network": (data.get("network") or "").strip(), "status": (data.get("status") or "รอขาย").strip(), "expiry_date": (data.get("expiry_date") or "").strip()}
            with db() as conn:
                ensure_master(conn, "networks", payload["network"])
                ensure_master(conn, "statuses", payload["status"])
                if payload["id"]:
                    conn.execute("UPDATE numbers SET sequence_no=:sequence_no, phone=:phone, sale_price=:sale_price, wholesale_price=:wholesale_price, network=:network, status=:status, expiry_date=:expiry_date, updated_at=CURRENT_TIMESTAMP WHERE id=:id", payload)
                else:
                    conn.execute("INSERT INTO numbers(sequence_no, phone, sale_price, wholesale_price, network, status, expiry_date) VALUES(:sequence_no, :phone, :sale_price, :wholesale_price, :network, :status, :expiry_date)", payload)
            return json_response(self, {"ok": True})
        if parsed.path == "/api/admin/delete":
            with db() as conn:
                conn.execute("DELETE FROM numbers WHERE id = ?", [self.read_json().get("id")])
            return json_response(self, {"ok": True})
        if parsed.path in {"/api/admin/statuses/save", "/api/admin/networks/save"}:
            try:
                save_master("statuses" if "statuses" in parsed.path else "networks", self.read_json())
                return json_response(self, {"ok": True})
            except Exception as exc:
                return json_response(self, {"error": str(exc)}, 400)
        if parsed.path in {"/api/admin/statuses/delete", "/api/admin/networks/delete"}:
            try:
                delete_master("statuses" if "statuses" in parsed.path else "networks", self.read_json().get("id"))
                return json_response(self, {"ok": True})
            except Exception as exc:
                return json_response(self, {"error": str(exc)}, 400)
        if parsed.path == "/api/admin/upload":
            replace_all = self.headers.get("X-Replace-All", "0") == "1"
            filename = self.headers.get("X-Filename", "upload.xlsx")
            tmp = DATA_DIR / f"upload-{secrets.token_hex(8)}.xlsx"
            tmp.write_bytes(self.read_body())
            try:
                summary = import_excel(tmp, filename, replace_all)
                return json_response(self, {"ok": True, **summary})
            except Exception as exc:
                with db() as conn:
                    detail = [{"row": 0, "phone": "", "status": "error", "message": traceback.format_exc()}]
                    conn.execute("INSERT INTO import_logs(filename, total_rows, imported_rows, duplicate_rows, error_rows, replaced_all, details) VALUES(?, 0, 0, 0, 1, ?, ?)", [filename, 1 if replace_all else 0, json.dumps(detail, ensure_ascii=False)])
                    log_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                return json_response(self, {"ok": False, "error": str(exc), "log_id": log_id}, 400)
            finally:
                tmp.unlink(missing_ok=True)
        return self.send_error(404)

    def export_rows(self, params):
        return query_numbers(params, limit=10000, admin_order=True)["items"]

    def template_xlsx(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "template"
        ws.append(["ลำดับ", "หมายเลขโทรศัพท์", "ราคาขาย", "ราคาส่ง", "เครือข่าย", "สถานะ", "วันหมดอายุ"])
        ws.append([1, "0812345678", 999, 500, "AIS", "รอขาย", "2026-12-31"])
        ws.append([2, "0899999999", 1999, 1000, "TRUE", "จองแล้ว", "2026-12-31"])
        bio = io.BytesIO()
        wb.save(bio)
        return file_response(self, bio.getvalue(), "template.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def export_xlsx(self, params):
        wb = Workbook()
        ws = wb.active
        ws.title = "RichBerD89"
        ws.append(["ลำดับ", "หมายเลขโทรศัพท์", "คะแนน", "ราคาขาย", "ราคาส่ง", "เครือข่าย", "สถานะ", "วันหมดอายุ"])
        for row in self.export_rows(params):
            ws.append([row["sequence_no"], row["phone"], row["analysis"]["score"], row["sale_price"], row["wholesale_price"], row["network"], row["status"], row["expiry_date"]])
        bio = io.BytesIO()
        wb.save(bio)
        return file_response(self, bio.getvalue(), "richber-export.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def export_pdf(self, params):
        bio = io.BytesIO()
        font = pdf_font_name()
        doc = SimpleDocTemplate(bio, pagesize=A4, rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        styles["Title"].fontName = font
        elements = [Paragraph("รายงานเบอร์ Rich Ber D 89", styles["Title"]), Spacer(1, 12)]
        data = [["ลำดับ", "เบอร์", "คะแนน", "ราคาขาย", "เครือข่าย", "สถานะ", "วันหมดอายุ"]]
        for row in self.export_rows(params)[:500]:
            data.append([row["sequence_no"], row["phone"], row["analysis"]["score"], f'{row["sale_price"]:,.0f}', row["network"], row["status"], row["expiry_date"]])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([("FONT", (0, 0), (-1, -1), font, 10), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0b2341")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#b8c2cc")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f7fb")])]))
        elements.append(table)
        doc.build(elements)
        return file_response(self, bio.getvalue(), "richber-export.pdf", "application/pdf")


if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", "8000"))
    host = os.environ.get("HOST", "127.0.0.1")
    server = ThreadingHTTPServer((host, port), App)
    print(f"Rich Ber D 89 running at http://{host}:{port}")
    server.serve_forever()
